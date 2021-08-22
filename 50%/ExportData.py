# System libraries
import os
from tensorflow.keras import models
from tensorflow.keras.layers import *
from tensorflow import keras
from tensorflow.keras.preprocessing.image import ImageDataGenerator
import openpyxl
import cv2
import numpy as np
from numba import jit
from numba import cuda
from tempfile import TemporaryFile

# User libraries
from modules.FileIO import *

def main():
    dim = (50,50)
    path = os.path.join(os.path.dirname(__file__))

    data_report_filename = 'fold.xlsx'
    mag = '40X'
    data_report_path = os.path.join(path, 'data-report', data_report_filename)
    
    workbook = openpyxl.load_workbook(data_report_path)
    sheet_1 = workbook['train']
    sheet_2 = workbook['validate']
    sheet_3 = workbook['test']
    
    train = get_data_from_data_report(sheet_1, dim, mag)
    # x_train = train[0]
    # y_train = train[1]
    # for y in y_train:
    #     try:
    #         new_y_train = np.append(new_y_train, int(float(y)))
    #     except:
    #         new_y_train = np.array(int(float(y)))
    # train_data = (x_train, new_y_train)
    save_numpy(path = path, 
               data = train, 
               mag = mag,
               subset = 'train')
    
    validate = get_data_from_data_report(sheet_2, dim, mag)
    # x_validate = validate[0]
    # y_validate = validate[1]
    # for y in y_validate:
    #     try:
    #         new_y_validate = np.append(new_y_validate, int(float(y)))
    #     except:
    #         new_y_validate = np.array(int(float(y)))
    # validate_data = (x_validate, new_y_validate)
    save_numpy(path = path, 
               data = validate, 
               mag = mag, 
               subset = 'validate')
    
    test = get_data_from_data_report(sheet_3, dim, mag)
    # x_test = test[0]
    # y_test = test[1]
    # for y in y_test:
    #     try:
    #         new_y_test = np.append(new_y_test, int(float(y)))
    #     except:
    #         new_y_test = np.array(int(float(y)))
    # test_data = (x_test, new_y_test)
    save_numpy(path = path, 
               data = test,  
               mag = mag,
               subset = 'test')
    cuda.profile_stop()

# @jit(forceobj = True)
def get_data_from_data_report(sheet, dim, magnification):
    x_dim = (sheet.max_row, dim[0], dim[1], 3)
    y_dim = (sheet.max_row)
    x_data = np.empty(x_dim, dtype = np.float)
    y_data = np.empty(y_dim, dtype = int)
    count = 0
    for row in sheet.iter_rows(min_row = 2, min_col = 1, max_row = sheet.max_row, max_col = 4):
        mag = row[1].value
        if mag != magnification:
            continue
        label = np.str(row[0].value) 
        img = cv2.imread(row[3].value)
        rs_img = cv2.resize(src = img, 
                            dsize = dim,
                            interpolation = cv2.INTER_AREA)
        x_data[count] = rs_img
        y_data[count] = label
        count += 1
    np.delete(x_data, 0)
    np.delete(y_data, 0)
    x_data = np.delete(x_data, range(count, sheet.max_row), axis = 0)
    y_data = np.delete(y_data, range(count, sheet.max_row), axis = 0)
    return x_data, y_data


if __name__ == "__main__":
    main()