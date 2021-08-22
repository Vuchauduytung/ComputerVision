# system library
import zipfile
import os
import numpy as np
import random
import sys
import cv2
from numba import jit
from numba import cuda
from timeit import default_timer as timer
from openpyxl import Workbook
import math

# User libraries
from modules.FileIO import *

def main():
    subset_ratio = (0.6, 0.2, 0.2)
    path = os.path.dirname(os.path.abspath(__file__))
    data_path = os.path.join(path, 'data\\BreaKHis_v1\\histology_slides\\breast')
    # data_array = get_data_array(data_path)
    data_structure_filename = 'breast.json'
    data_structure_path = os.path.join(path, 'drivers\\data-structure', data_structure_filename)
    data_structure = json2dict(data_structure_path)
    data_direct = get_data_direct(data_path, data_structure)
    data_direct_array = get_data_direct_array(data_path, data_direct)

    train_data_path, validate_data_path, test_data_path = get_subset_path(data_direct, data_direct_array, subset_ratio)  
    
    workbook = Workbook()
    data_report_filename = 'fold.xlsx'
    data_report_path = os.path.join(path, 'data-report', data_report_filename)
    sheet_1 = workbook.active
    sheet_1.title = 'train'
    sheet_2 = workbook.create_sheet(title = 'validate')
    sheet_3 = workbook.create_sheet(title = 'test')
    update_sheet(sheet_1, train_data_path)         
    update_sheet(sheet_2, validate_data_path)         
    update_sheet(sheet_3, test_data_path)                  
    workbook.save(filename = data_report_path)
    
def get_data_direct(data_path, data_structure):
    direct = np.array([])
    for key, value in data_structure.items():
        if isinstance(value, dict):
            sub_data_path = os.path.join(data_path, key)
            sub_direct = get_data_direct(sub_data_path, value)
            direct = np.append(direct, sub_direct)
        elif isinstance(value, list):
            sub_direct = np.array([])
            for val in value:
                if isinstance(val, str):
                    sub_direct = np.append(sub_direct, val)        
            for sub_dir in sub_direct:
                directory = os.path.join(data_path, key, sub_dir)
                direct = np.append(direct, directory)
    return direct

def get_data_direct_array(data_path, data_direct):
    data_direct_array = np.array([])
    for direct in data_direct:
        data_directory = os.path.join(data_path, direct)
        for (root,_,files) in os.walk(data_directory, topdown= True):
            for file in files:
                if file.endswith('.png'):
                    data_dir = os.path.join(root, file)
                    data_direct_array = np.append(data_direct_array, data_dir)
    return data_direct_array

def get_subset_path(data_direct, data_direct_array, subset_ratio):

    data_group = {}
    for data_dir in data_direct:
        direct = np.array([])
        for data_direct_ in data_direct_array:
            if data_dir in data_direct_:
                direct = np.append(direct, data_direct_)
        data_group[data_dir] = direct
    
    train_data_path = {}
    validate_data_path = {}
    test_data_path = {}
    for data_dir, direct in data_group.items():
        np.random.shuffle(direct)
        train_count = math.floor(len(direct)*subset_ratio[0])
        if train_count == 0: 
            train_count = len(direct)
            validate_count = 0
            test_count = 0
        else: 
            validate_count = math.floor(len(direct)*subset_ratio[1])
            test_count = math.floor(len(direct)*subset_ratio[2])
        train_index = range(train_count)
        validate_index = range(train_count, train_count + validate_count)
        test_index = range(train_count + validate_count, train_count + validate_count + test_count)
        train_path = direct[train_index]
        validate_path = direct[validate_index]
        test_path = direct[test_index]  
        train_data_path.update({data_dir : train_path})   
        validate_data_path.update({data_dir : validate_path})   
        test_data_path.update({data_dir : test_path})   
    
    return train_data_path, validate_data_path, test_data_path

def update_sheet(sheet, data_path):
    sheet['A1'] = 'fold'
    sheet['B1'] = 'mag'
    sheet['C1'] = 'parent'
    sheet['D1'] = 'file path'

    fold = {
        '0' : 'ductal_carcinoma',
        '1' : 'lobular_carcinoma',
        '2' : 'mucinous_carcinoma',
        '3' : 'papillary_carcinoma'
    }
    

    for direc_path in data_path.values():
        try:
            data_path_full = np.append(data_path_full, direc_path)
        except NameError:
            data_path_full = np.array(direc_path)
    count = 1
    for key, class_name in fold.items(): 
        mag = ''
        old_mag = []      
                     
        for direct in data_path_full:
            if class_name not in direct.split('\\'):
                continue
            path_1 = os.path.split(direct)[0]
            mag = os.path.split(path_1)[1]
            parent = ''
            old_parent = []
            if mag in old_mag:
                continue
            
            for direct_ in data_path_full:
                # parent = os.path.split(path_2)[1]
                if class_name not in direct_.split('\\'):
                    continue
                if mag not in direct_.split('\\'):
                    continue
                path_1_ = os.path.split(direct_)[0]
                path_2_= os.path.split(path_1_)[0]
                parent = os.path.split(path_2_)[1]
                if parent in old_parent:
                    continue   
                        
                for direct__ in data_path_full:
                    if class_name not in direct__.split('\\'):
                        continue
                    if mag not in direct__.split('\\'):
                        continue
                    if parent not in direct__.split('\\'):
                        continue
                    if mag not in old_mag:
                        old_mag.append(mag)
                    if parent not in old_parent:
                        old_parent.append(parent)     
                    count += 1
                    sheet.cell(column = 1, 
                                row = count, 
                                value = key)
                    sheet.cell(column = 2, 
                                row = count, 
                                value = mag)
                    sheet.cell(column = 3, 
                                row = count, 
                                value = parent) 
                    sheet.cell(column = 4, 
                                row = count, 
                                value = direct__)            
                # old_parent = []
        # old_mag = []            

if __name__ == '__main__':
    main()